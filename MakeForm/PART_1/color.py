import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import time


def color(sheet_names):
    # 代码备注请参考 函数track(sheet_names)，大部分内容极其相似
    special_match = {'commom': ['Study', 'SasText', 'SubjectNo', 'SubjectName', 'SubjectStatus', 'VersionNo', 'SiteNo', 'SiteName', 'Instance', 'Block', 'DataPage','Line'],
                 'LB11': ['SubjectNo', 'Instance', 'Block', 'DataPage','Line'],
                 'LB12': ['SubjectNo', 'Instance', 'Block', 'DataPage','Line']}
    ignore_value = {'commom': ['LastModifiedDate', 'Study', 'SasText', 'SubjectNo', 'SubjectName', 'SubjectStatus', 'VersionNo', 'SiteNo', 'SiteName', 'Instance', 'Block', 'DataPage','Line']}

    dif_dict = {}

    xie = pd.ExcelWriter(save)
    for each in sheet_names:
        if each in special_match.keys():
            key_col = special_match[each]
        else:
            key_col = special_match['commom']
        if each in ignore_value.keys():
            pass_value = ignore_value[each]
        else:
            pass_value = ignore_value['commom']
        sheet_new = pd.read_excel(new, sheet_name=each, keep_default_na=False)
        sheet_old = pd.read_excel(old, sheet_name=each, keep_default_na=False).rename(columns={col: col+'_old' for col in key_col})
        res = pd.merge(sheet_new, sheet_old, how='left', left_on=key_col, right_on=[k+'_old' for k in key_col])
        
        dif_cells = [[], []]
        for row in range(res.shape[0]):
            for col in range(1, sheet_new.shape[1]):  # 第一列不参与，作RUNNING DATE的识别标
                if list(sheet_new.columns)[col] in pass_value:
                    pass
                else:
                    new_cell = res.iloc[row, col]
                    old_cell = res.iloc[row, col+sheet_new.shape[1]]
                    if new_cell != old_cell:
                        dif_cells[0].append(row+2)
                        dif_cells[1].append(col+1)
                    else:
                        pass
        dif_dict[each] = dif_cells
        res = res.iloc[:, list(range(sheet_new.shape[1]))+list(range(sheet_new.shape[1]*2, res.shape[1]))]  # 删除右边的旧信息，保留左边的新信息
        res.rename(columns={col: col.replace('_x', '') for col in list(res.columns)}, inplace=True)
        res.to_excel(xie, sheet_name=each, index=False)
    xie.close()

    wb = openpyxl.load_workbook(save)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        color_cells = dif_dict[sheet]
        rows = sorted(set(color_cells[0]))
        for row in rows:
            for col in range(1, ws.max_column+1):
                ws.cell(row, col).fill = PatternFill('solid', fgColor='98F5FF')
        for idx in range(len(color_cells[0])):
            row = color_cells[0][idx]
            col = color_cells[1][idx]
            ws.cell(row, col).fill = PatternFill('solid', fgColor='FFFF00')
    wb.save(save)
    wb.close()


def track(sheet_names):
    # 通过ignore_value，指定哪些列是前后对应的依据
    special_match = {'commom': ['受试者编号', '数据页', '行号'],
                     'MH1': ['受试者编号', '数据页']}
    # 通过自定义delete_value，确保新表shape[1]和旧表shape[1]主结构的大小一致且列对应
    delete_value = {'commom': ['项目名称','研究中心名称','受试者姓名缩写','数据节','数据页标识符OID','数据块'],
                    'MH1': ['行号', '项目名称','研究中心名称','受试者姓名缩写','数据节','数据页标识符OID','数据块']}
    # 通过ignore_value，指定哪些列不需要参与新旧差异比对
    ignore_value = {'commom': ['受试者编号','数据页','行号','编码字典'],}
    dif_dict = {} # 表单名：差异单元格坐标

    xie = pd.ExcelWriter(save)
    for each in sheet_names:
        if each in special_match.keys():
            key_col = special_match[each]
        else:
            key_col = special_match['commom']
        if each in delete_value.keys():
            d_value = delete_value[each]
        else:
            d_value = delete_value['commom']
        if each in ignore_value.keys():
            pass_value = ignore_value[each]
        else:
            pass_value = ignore_value['commom']

        sheet_new = pd.read_excel(new, sheet_name=each, keep_default_na=False).drop(columns=d_value)
        for col in list(sheet_new.columns):
            sheet_new[col] = sheet_new[col].astype(str)
        
        sheet_old = pd.read_excel(old, sheet_name=each.lower(), keep_default_na=False).rename(columns={col+' ': col+'_old' for col in key_col})  # lower\空格字符 后续可能会删除
        for col in list(sheet_old.columns):
            sheet_old[col] = sheet_old[col].astype(str)

        for row in range(sheet_old.shape[0]):
            for col in range(sheet_old.shape[1]):
                sheet_old.iloc[row, col] = str(sheet_old.iloc[row, col]).strip(' ')

        for row in range(sheet_new.shape[0]):
            for col in range(sheet_new.shape[1]):
                sheet_new.iloc[row, col] = str(sheet_new.iloc[row, col]).strip(' ')

        res = pd.merge(sheet_new, sheet_old, how='left', left_on=key_col, right_on=[k+'_old' for k in key_col])
        dif_cells = [[], []]
        for row in range(res.shape[0]):
            for col in range(1, sheet_new.shape[1]):  # 第一列不参与，作RUNNING DATE的识别标
                if list(sheet_new.columns)[col] in pass_value:
                    pass
                else:
                    new_cell = res.iloc[row, col]
                    old_cell = res.iloc[row, col+sheet_new.shape[1]]
                    if new_cell != old_cell:
                        dif_cells[0].append(row+2)
                        dif_cells[1].append(col+1)
                    else:
                        pass
        dif_dict[each.lower()] = dif_cells

        res = res.iloc[:, list(range(sheet_new.shape[1]))+list(range(sheet_new.shape[1]*2, res.shape[1]))]  # 删除右边的旧信息，保留左边的新信息

        res.to_excel(xie, sheet_name=each.lower(), index=False)  # lower以后可能删除
    xie.close()

    wb = openpyxl.load_workbook(save)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        color_cells = dif_dict[sheet]
        rows = sorted(set(color_cells[0]))
        for row in rows:
            for col in range(1, ws.max_column+1):
                ws.cell(row, col).fill = PatternFill('solid', fgColor='98F5FF')
        for idx in range(len(color_cells[0])):
            row = color_cells[0][idx]
            col = color_cells[1][idx]
            ws.cell(row, col).fill = PatternFill('solid', fgColor='FFFF00')
    wb.save(save)
    wb.close()


def add_rd(res_file, running_date_name = '日期'):  # 自定义running_date_name的字符
    wb = openpyxl.load_workbook(res_file)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        head_cols = [ws.cell(1, col).value.strip(' ') for col in range(1, ws.max_column+1)]
        if running_date_name in head_cols:  # 如果有running date，比差异
            date_idx = head_cols.index(running_date_name) + 1
            for row in range(2, ws.max_row+1):
                if ws.cell(row, 1).fill.fgColor.rgb != '00000000':
                    ws.cell(row, date_idx).value = today
                else:
                    pass
        else:
            ws.insert_cols(len(head_cols)+1)  # 如果没有，插入列，所有行加上running date，均为当此运行日期
            ws.cell(1, len(head_cols)+1).value = running_date_name
            date_idx = len(head_cols) + 1
            for row in range(2, ws.max_row+1):
                ws.cell(row, date_idx).value = today
    wb.save(res_file)
    wb.close()


def main(cmd):
    global new, old, save, today
    st = time.localtime()
    today = '{}/{}/{}'.format(st[0], st[1], st[2])
    win = tk.Tk()
    win.withdraw()
    new = filedialog.askopenfilename(title='请选择最近的源文件')
    old = filedialog.askopenfilename(title='请选择上次的处理文件')
    save = filedialog.askdirectory(title='请选择保存路径')+'/差异标色结果.xlsx'
    """
    1：源文件
    2:处理后文件
    """
    if cmd == 1:
        sheets_new = list(pd.ExcelFile(new).sheet_names)
        sheets_old = list(pd.ExcelFile(old).sheet_names)
        sheets = [ws for ws in sheets_new if ws in sheets_old]
        if 'TOC' in sheets:
            sheets.remove('TOC')
        color(sheets)
        add_rd(save)
    elif cmd == 2:
        sheets_new = list(pd.ExcelFile(new).sheet_names)
        sheets_old = list(pd.ExcelFile(old).sheet_names)
        sheets = [ws for ws in sheets_new if ws.lower() in sheets_old]  # lower以后可能删除
        if 'TOC' in sheets:
            sheets.remove('TOC')
        track(sheets)
        add_rd(save)
    else:
        pass


main(1)
